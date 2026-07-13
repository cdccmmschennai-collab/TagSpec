"""Helpers to build in-memory test workbooks with rich structure."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation


def build_basic_workbook(
    *,
    header_row: int = 1,
    headers: tuple[str, str, str] = (
        "TAG NUMBER REV-1",
        "EQUIPMENT DESCRIPTION",
        "ADDITIONAL INFORMATION",
    ),
    rows: list[tuple[str, str, str]] | None = None,
    lead_columns: int = 0,
    metadata_rows: int = 0,
) -> bytes:
    """Return .xlsx bytes with the three headers plus optional data rows.

    ``lead_columns`` inserts filler columns before the tag column so header
    detection cannot rely on fixed positions.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "TAGS"

    for i in range(metadata_rows):
        ws.cell(row=i + 1, column=1).value = f"PROJECT META {i}"

    hr = header_row
    for offset, header in enumerate(headers):
        ws.cell(row=hr, column=lead_columns + 1 + offset).value = header

    for r_offset, (tag, equip, addl) in enumerate(rows or [], start=1):
        ws.cell(row=hr + r_offset, column=lead_columns + 1).value = tag
        ws.cell(row=hr + r_offset, column=lead_columns + 2).value = equip
        ws.cell(row=hr + r_offset, column=lead_columns + 3).value = addl

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_rich_workbook() -> bytes:
    """Workbook with multiple sheets, formulas, merges, freeze, hidden, etc."""
    wb = Workbook()

    data = wb.active
    data.title = "TAGS"
    data["A1"] = "TAG NUMBER REV-1"
    data["B1"] = "EQUIPMENT DESCRIPTION"
    data["C1"] = "ADDITIONAL INFORMATION"
    data["D1"] = "QTY"
    data["A2"] = "P-101"
    data["B2"] = "BALL VALVE"
    data["C2"] = None
    data["D2"] = 5
    data["A3"] = "P-102"
    data["B3"] = "BALL VALVE"
    data["C3"] = "PREEXISTING:VALUE"
    data["D3"] = 3
    data["E2"] = "=D2*2"  # formula preserved
    data.column_dimensions["A"].width = 30
    data.row_dimensions[2].height = 42
    data.freeze_panes = "B2"

    other = wb.create_sheet("NOTES")
    other["A1"] = "Keep me"
    other["B1"] = "=1+2"
    other.merge_cells("A3:C3")
    other["A3"] = "MERGED HEADER"
    other.column_dimensions["Z"].hidden = True
    dv = DataValidation(type="list", formula1='"YES,NO"')
    other.add_data_validation(dv)
    dv.add("A5")

    hidden = wb.create_sheet("HIDDEN_SHEET")
    hidden["A1"] = "secret"
    hidden.sheet_state = "hidden"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
