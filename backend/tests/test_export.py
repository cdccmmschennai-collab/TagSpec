"""Excel export + workbook preservation regression tests (Phase 10)."""

from __future__ import annotations

import hashlib
import uuid

from openpyxl import load_workbook
from sqlalchemy import select

from app.core.errors import ConflictError
from app.models.export import ExportHistory
from app.models.tag_row import ImportedTagRow
from app.models.workbook import WorkbookJob
from app.services import export_service
from tests.helpers import create_equipment
from tests.workbook_factory import build_rich_workbook


def _upload_rich(client, headers):
    content = build_rich_workbook()
    resp = client.post(
        "/api/v1/workbooks",
        headers=headers,
        files={"file": ("rich.xlsx", content, "application/octet-stream")},
        data={"project_code": "P389", "revision": "REV1"},
    )
    assert resp.status_code == 201, resp.text
    return resp.json()["job"]


def _complete_p101(client, auth, job_id):
    tags = client.get(
        f"/api/v1/workbooks/{job_id}/tags",
        headers=auth["editor"],
        params={"equipment": "BALL VALVE"},
    ).json()
    tag = next(t for t in tags if t["tag_number"] == "P-101")
    claim = client.post(f"/api/v1/workbooks/tags/{tag['id']}/claim", headers=auth["editor"]).json()
    resp = client.post(
        f"/api/v1/workbooks/tags/{tag['id']}/complete",
        headers=auth["editor"],
        json={"values": {"BODY": "A105N", "SEAT": "F316L"}, "row_version": claim["tag"]["row_version"]},
    )
    assert resp.status_code == 200, resp.text
    return resp.json()["generated_additional_information"]


def test_export_updates_only_target_cell_and_preserves_workbook(client, auth, db):
    create_equipment(client, auth["admin"], "BALL VALVE", attrs=("BODY", "SEAT"))
    job = _upload_rich(client, auth["supervisor"])
    client.post(f"/api/v1/workbooks/{job['id']}/import", headers=auth["supervisor"])
    generated = _complete_p101(client, auth, job["id"])

    job_row = db.get(WorkbookJob, uuid.UUID(job["id"]))
    original_path = job_row.original_file_path
    original_hash_before = hashlib.sha256(open(original_path, "rb").read()).hexdigest()

    resp = client.post(f"/api/v1/workbooks/{job['id']}/export", headers=auth["supervisor"])
    assert resp.status_code == 200, resp.text
    export = resp.json()
    assert export["status"] == "SUCCESS"
    assert export["row_count_written"] == 1

    # 1) Original file untouched.
    original_hash_after = hashlib.sha256(open(original_path, "rb").read()).hexdigest()
    assert original_hash_before == original_hash_after

    # 2) Exported workbook reopens and target cell updated; others preserved.
    export_record = db.scalars(select(ExportHistory)).first()
    out_path = export_record.output_file_path

    wb = load_workbook(out_path)  # if this fails, workbook did not reopen
    ws = wb["TAGS"]
    assert ws["C2"].value == generated                 # target updated
    assert ws["C3"].value == "PREEXISTING:VALUE"        # existing untouched
    assert ws["D2"].value == 5                          # unrelated value kept
    assert ws["E2"].value == "=D2*2"                    # formula preserved
    assert ws.freeze_panes == "B2"                      # freeze preserved
    assert abs(ws.column_dimensions["A"].width - 30) < 0.01
    assert abs(ws.row_dimensions[2].height - 42) < 0.01

    notes = wb["NOTES"]
    assert "A3:C3" in [str(r) for r in notes.merged_cells.ranges]
    assert notes["B1"].value == "=1+2"
    assert notes.column_dimensions["Z"].hidden is True
    assert len(notes.data_validations.dataValidation) == 1

    assert wb["HIDDEN_SHEET"].sheet_state == "hidden"
    assert wb.sheetnames == ["TAGS", "NOTES", "HIDDEN_SHEET"]

    # exported rows marked EXPORTED
    db.expire_all()
    exported_row = db.scalars(
        select(ImportedTagRow).where(ImportedTagRow.tag_number == "P-101")
    ).first()
    assert exported_row.status == "EXPORTED"


def test_duplicate_concurrent_export_blocked(client, auth, db):
    create_equipment(client, auth["admin"], "BALL VALVE", attrs=("BODY", "SEAT"))
    job = _upload_rich(client, auth["supervisor"])
    client.post(f"/api/v1/workbooks/{job['id']}/import", headers=auth["supervisor"])
    _complete_p101(client, auth, job["id"])

    job_uuid = uuid.UUID(job["id"])
    lock = export_service._job_lock(job_uuid)
    lock.acquire()
    try:
        try:
            export_service.export_job(db, job_uuid, None)
            raise AssertionError("expected EXPORT_IN_PROGRESS conflict")
        except ConflictError as exc:
            assert exc.code == "EXPORT_IN_PROGRESS"
    finally:
        lock.release()


def test_export_history_recorded(client, auth):
    create_equipment(client, auth["admin"], "BALL VALVE", attrs=("BODY", "SEAT"))
    job = _upload_rich(client, auth["supervisor"])
    client.post(f"/api/v1/workbooks/{job['id']}/import", headers=auth["supervisor"])
    _complete_p101(client, auth, job["id"])
    client.post(f"/api/v1/workbooks/{job['id']}/export", headers=auth["supervisor"])
    history = client.get(f"/api/v1/workbooks/{job['id']}/exports", headers=auth["supervisor"]).json()
    assert len(history) == 1
    assert history[0]["status"] == "SUCCESS"
    assert history[0]["file_hash"]
