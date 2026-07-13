"""Workbook upload, secure storage, and header/structure detection."""

from __future__ import annotations

import hashlib
import re
import uuid
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.errors import ConflictError, NotFoundError, ValidationAppError
from app.core.normalization import normalize_header
from app.models.enums import JobStatus
from app.models.workbook import WorkbookJob

# Required headers (normalized form) and the job-column they map to.
REQUIRED_HEADERS = {
    "TAG NUMBER REV-1": "tag_column_number",
    "EQUIPMENT DESCRIPTION": "equipment_column_number",
    "ADDITIONAL INFORMATION": "additional_info_column_number",
}

XLSX_MAGIC = b"PK\x03\x04"  # xlsx is a zip archive
_SAFE_NAME_RE = re.compile(r"[^A-Za-z0-9._-]")


@dataclass
class DetectionResult:
    sheet_name: str
    header_row_number: int
    tag_column_number: int
    equipment_column_number: int
    additional_info_column_number: int
    detected_headers: dict[str, int] = field(default_factory=dict)


def sanitize_filename(name: str) -> str:
    """Strip path components and unsafe characters; keep a plain basename."""
    base = Path(name).name  # drops directories / traversal
    base = base.replace("\x00", "")
    cleaned = _SAFE_NAME_RE.sub("_", base).strip("._") or "workbook"
    if not cleaned.lower().endswith(".xlsx"):
        cleaned += ".xlsx"
    return cleaned[:200]


def _stored_path(stored_filename: str) -> Path:
    upload_dir = Path(settings.upload_dir).resolve()
    upload_dir.mkdir(parents=True, exist_ok=True)
    target = (upload_dir / stored_filename).resolve()
    # Path-traversal guard: resolved target must stay inside upload_dir.
    if upload_dir not in target.parents and target != upload_dir:
        raise ValidationAppError("Invalid stored path", code="PATH_TRAVERSAL")
    return target


def detect_structure(path: Path) -> DetectionResult:
    """Find the worksheet/row/columns for the three required headers."""
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises many types on malformed files
        raise ValidationAppError(
            f"File is not a readable .xlsx workbook: {exc}",
            code="MALFORMED_WORKBOOK",
        ) from exc

    max_rows = settings.header_search_rows
    try:
        for ws in wb.worksheets:
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1
            ):
                found: dict[str, int] = {}
                for col_idx, value in enumerate(row, start=1):
                    header = normalize_header(value)
                    if header in REQUIRED_HEADERS and header not in found:
                        found[header] = col_idx
                if all(h in found for h in REQUIRED_HEADERS):
                    return DetectionResult(
                        sheet_name=ws.title,
                        header_row_number=row_idx,
                        tag_column_number=found["TAG NUMBER REV-1"],
                        equipment_column_number=found["EQUIPMENT DESCRIPTION"],
                        additional_info_column_number=found["ADDITIONAL INFORMATION"],
                        detected_headers=dict(found),
                    )
    finally:
        wb.close()

    raise ValidationAppError(
        "Required columns not found. Expected headers: "
        + ", ".join(REQUIRED_HEADERS),
        code="MISSING_REQUIRED_COLUMNS",
    )


def _validate_upload_bytes(content: bytes, original_filename: str) -> None:
    if not original_filename.lower().endswith(".xlsx"):
        raise ValidationAppError(
            "Only .xlsx files are supported", code="UNSUPPORTED_EXTENSION"
        )
    if len(content) == 0:
        raise ValidationAppError("Uploaded file is empty", code="EMPTY_FILE")
    if len(content) > settings.max_upload_size_bytes:
        raise ValidationAppError(
            f"File exceeds the {settings.max_upload_size_mb} MB limit",
            code="FILE_TOO_LARGE",
        )
    if not content.startswith(XLSX_MAGIC):
        raise ValidationAppError(
            "File is not a valid .xlsx workbook (bad signature)",
            code="INVALID_WORKBOOK_SIGNATURE",
        )


def find_job_by_hash(db: Session, file_hash: str) -> WorkbookJob | None:
    return db.scalars(
        select(WorkbookJob).where(WorkbookJob.file_hash == file_hash)
    ).first()


def create_job_from_upload(
    db: Session,
    *,
    content: bytes,
    original_filename: str,
    job_name: str | None,
    project_code: str | None,
    revision: str | None,
    uploaded_by: uuid.UUID | None,
    allow_duplicate: bool = False,
) -> tuple[WorkbookJob, DetectionResult, bool]:
    """Validate, store and register a workbook. Returns (job, detection, is_dup)."""
    _validate_upload_bytes(content, original_filename)

    file_hash = hashlib.sha256(content).hexdigest()
    existing = find_job_by_hash(db, file_hash)
    if existing is not None and not allow_duplicate:
        detection = DetectionResult(
            sheet_name=existing.sheet_name or "",
            header_row_number=existing.header_row_number or 0,
            tag_column_number=existing.tag_column_number or 0,
            equipment_column_number=existing.equipment_column_number or 0,
            additional_info_column_number=existing.additional_info_column_number or 0,
        )
        return existing, detection, True

    safe_original = sanitize_filename(original_filename)
    stored_filename = f"{uuid.uuid4().hex}_{safe_original}"
    target = _stored_path(stored_filename)
    target.write_bytes(content)

    try:
        detection = detect_structure(target)
    except ValidationAppError:
        target.unlink(missing_ok=True)
        raise

    now = datetime.now(UTC)
    job = WorkbookJob(
        job_name=(job_name or safe_original).strip(),
        project_code=project_code,
        revision=revision,
        original_filename=original_filename[:512],
        stored_filename=stored_filename,
        original_file_path=str(target),
        file_hash=file_hash,
        file_size=len(content),
        sheet_name=detection.sheet_name,
        header_row_number=detection.header_row_number,
        tag_column_number=detection.tag_column_number,
        equipment_column_number=detection.equipment_column_number,
        additional_info_column_number=detection.additional_info_column_number,
        status=JobStatus.READY.value,
        uploaded_by=uploaded_by,
        uploaded_at=now,
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    return job, detection, False


def get_job(db: Session, job_id: uuid.UUID) -> WorkbookJob:
    job = db.get(WorkbookJob, job_id)
    if job is None:
        raise NotFoundError("Workbook job not found")
    return job


def list_jobs(db: Session) -> list[WorkbookJob]:
    return list(
        db.scalars(select(WorkbookJob).order_by(WorkbookJob.created_at.desc())).all()
    )


def delete_job(db: Session, job_id: uuid.UUID, *, delete_file: bool = True) -> None:
    from app.models.tag_row import ImportedTagRow

    job = get_job(db, job_id)
    # Safety: refuse to delete once any completed/exported work exists.
    from app.models.enums import TagStatus

    locked = db.scalars(
        select(ImportedTagRow).where(
            ImportedTagRow.workbook_job_id == job.id,
            ImportedTagRow.status.in_(
                [TagStatus.COMPLETED.value, TagStatus.EXPORTED.value]
            ),
        )
    ).first()
    if locked is not None:
        raise ConflictError(
            "Cannot delete a job that has completed or exported tags",
            code="JOB_HAS_WORK",
        )
    if delete_file and job.original_file_path:
        Path(job.original_file_path).unlink(missing_ok=True)
    db.delete(job)
    db.commit()
