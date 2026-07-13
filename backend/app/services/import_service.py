"""Import tag rows from a validated workbook job into PostgreSQL."""

from __future__ import annotations

import re
import uuid
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.errors import ConflictError, ValidationAppError
from app.core.normalization import (
    normalize_equipment_description,
    normalize_tag_number,
)
from app.models.enums import JobStatus, TagStatus
from app.models.tag_row import ImportedTagRow
from app.models.workbook import WorkbookJob
from app.schemas.workbook import EquipmentDescriptionCount, ImportStats
from app.services import equipment_service


def _cell(row: tuple, col_number: int) -> object:
    idx = col_number - 1
    if 0 <= idx < len(row):
        return row[idx]
    return None


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


# A bare number: integer or decimal, optional sign. Used to spot metadata rows.
_NUMERIC_RE = re.compile(r"[+-]?\d+(?:\.\d+)?")


def _looks_like_metadata_value(value: str) -> bool:
    """True when ``value`` is a purely numeric cell.

    SAP-style exports place a metadata band immediately under the header row —
    field codes, field lengths and a field column-number row. In that band the
    EQUIPMENT DESCRIPTION column holds a bare number (e.g. ``6``), whereas a
    real equipment description is always textual (``BALL VALVE``). Rejecting
    purely numeric equipment values drops the metadata row without touching any
    legitimate tag: descriptions that merely *contain* digits (``3 WAY VALVE``)
    are unaffected.
    """
    return bool(_NUMERIC_RE.fullmatch(value.strip()))


def import_rows(db: Session, job: WorkbookJob, *, rerun: bool = False) -> ImportStats:
    """Read the detected sheet and (re)create tag rows.

    On rerun, existing rows are re-matched but claimed/completed work is
    preserved (only AVAILABLE / TEMPLATE_MISSING rows are refreshed).
    """
    if not job.sheet_name or not job.header_row_number:
        raise ValidationAppError("Workbook has not been validated", code="JOB_NOT_VALIDATED")

    existing = db.scalars(
        select(ImportedTagRow).where(ImportedTagRow.workbook_job_id == job.id)
    ).all()
    if existing and not rerun:
        raise ConflictError("Rows already imported for this job", code="ALREADY_IMPORTED")
    existing_by_row = {r.excel_row_number: r for r in existing}

    path = Path(job.original_file_path)
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb[job.sheet_name]

        total_inspected = 0
        tag_rows = 0
        equipment_norms: set[str] = set()
        matched = 0
        unmatched = 0
        available = 0
        template_missing = 0
        existing_in_excel = 0
        norm_tag_counter: Counter[str] = Counter()

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=job.header_row_number + 1, values_only=True), start=job.header_row_number + 1
        ):
            total_inspected += 1
            tag_raw = _text(_cell(row, job.tag_column_number))
            equip_raw = _text(_cell(row, job.equipment_column_number))
            addl_raw = _text(_cell(row, job.additional_info_column_number))

            # Only real tag records: both tag and equipment must be present.
            if not tag_raw or not equip_raw:
                continue

            # Skip the SAP metadata band (field column-number / length rows) that
            # sits just under the header: a real equipment description is textual,
            # so a purely numeric value in that column is never a tag record.
            if _looks_like_metadata_value(equip_raw):
                continue

            tag_rows += 1
            norm_tag = normalize_tag_number(tag_raw)
            norm_equip = normalize_equipment_description(equip_raw)
            equipment_norms.add(norm_equip)
            norm_tag_counter[norm_tag] += 1

            template = equipment_service.match_equipment(db, norm_equip)
            if template is not None:
                matched += 1
            else:
                unmatched += 1

            # Decide status.
            if addl_raw:
                status = TagStatus.EXISTING_IN_EXCEL.value
                existing_in_excel += 1
            elif template is not None:
                status = TagStatus.AVAILABLE.value
                available += 1
            else:
                status = TagStatus.TEMPLATE_MISSING.value
                template_missing += 1

            row_model = existing_by_row.get(row_idx)
            if row_model is None:
                row_model = ImportedTagRow(
                    workbook_job_id=job.id,
                    sheet_name=job.sheet_name,
                    excel_row_number=row_idx,
                )
                db.add(row_model)

            # Preserve in-flight/completed work on rerun.
            protected = {
                TagStatus.CLAIMED.value,
                TagStatus.DRAFT.value,
                TagStatus.COMPLETED.value,
                TagStatus.REVIEW_REQUIRED.value,
                TagStatus.REVIEWED.value,
                TagStatus.EXPORTED.value,
            }
            row_model.tag_number = tag_raw
            row_model.normalized_tag_number = norm_tag
            row_model.equipment_description = equip_raw
            row_model.normalized_equipment_description = norm_equip
            row_model.equipment_template_id = template.id if template else None
            row_model.original_additional_information = addl_raw or None
            if row_model.status not in protected:
                row_model.status = status

        db.commit()
    finally:
        wb.close()

    duplicates = sorted(t for t, c in norm_tag_counter.items() if c > 1)

    if job.status in {JobStatus.UPLOADED.value, JobStatus.READY.value, JobStatus.VALIDATING.value}:
        job.status = JobStatus.IN_PROGRESS.value
        db.commit()

    return ImportStats(
        total_rows_inspected=total_inspected,
        total_tag_rows=tag_rows,
        equipment_description_count=len(equipment_norms),
        matched_equipment_count=matched,
        unmatched_equipment_count=unmatched,
        available_tags=available,
        template_missing_tags=template_missing,
        existing_in_excel_tags=existing_in_excel,
        duplicate_tag_numbers=duplicates,
    )


def equipment_description_counts(db: Session, job_id: uuid.UUID) -> list[EquipmentDescriptionCount]:
    rows = db.scalars(
        select(ImportedTagRow).where(ImportedTagRow.workbook_job_id == job_id)
    ).all()
    buckets: dict[str, dict] = defaultdict(
        lambda: {
            "equipment_description": "",
            "equipment_template_id": None,
            "total": 0,
            "available": 0,
            "in_progress": 0,
            "completed": 0,
            "existing_in_excel": 0,
            "template_missing": 0,
        }
    )
    for r in rows:
        key = r.normalized_equipment_description or ""
        b = buckets[key]
        b["equipment_description"] = r.equipment_description or key
        b["equipment_template_id"] = r.equipment_template_id
        b["total"] += 1
        if r.status == TagStatus.AVAILABLE.value:
            b["available"] += 1
        elif r.status in {TagStatus.CLAIMED.value, TagStatus.DRAFT.value}:
            b["in_progress"] += 1
        elif r.status in {TagStatus.COMPLETED.value, TagStatus.REVIEWED.value, TagStatus.EXPORTED.value}:
            b["completed"] += 1
        elif r.status == TagStatus.EXISTING_IN_EXCEL.value:
            b["existing_in_excel"] += 1
        elif r.status == TagStatus.TEMPLATE_MISSING.value:
            b["template_missing"] += 1

    result = []
    for norm, b in sorted(buckets.items()):
        result.append(
            EquipmentDescriptionCount(
                normalized_equipment_description=norm,
                equipment_description=b["equipment_description"],
                equipment_template_id=b["equipment_template_id"],
                total=b["total"],
                available=b["available"],
                in_progress=b["in_progress"],
                completed=b["completed"],
                existing_in_excel=b["existing_in_excel"],
                template_missing=b["template_missing"],
            )
        )
    return result


def unmatched_descriptions(db: Session, job_id: uuid.UUID) -> list[str]:
    rows = db.scalars(
        select(ImportedTagRow.normalized_equipment_description).where(
            ImportedTagRow.workbook_job_id == job_id,
            ImportedTagRow.equipment_template_id.is_(None),
        )
    ).all()
    return sorted({r for r in rows if r})


def assign_description_to_template(
    db: Session,
    job_id: uuid.UUID,
    normalized_description: str,
    template_id: uuid.UUID,
) -> int:
    """Point all rows with the given description at a template and refresh status."""
    template = equipment_service.get_equipment(db, template_id)
    rows = db.scalars(
        select(ImportedTagRow).where(
            ImportedTagRow.workbook_job_id == job_id,
            ImportedTagRow.normalized_equipment_description == normalized_description,
        )
    ).all()
    updated = 0
    for r in rows:
        r.equipment_template_id = template.id
        if r.status == TagStatus.TEMPLATE_MISSING.value and not r.original_additional_information:
            r.status = TagStatus.AVAILABLE.value
        updated += 1
    db.commit()
    return updated
