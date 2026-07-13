"""Server-side Excel export. Updates ONLY Additional Information cells."""

from __future__ import annotations

import hashlib
import os
import tempfile
import threading
import uuid
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.errors import ConflictError, NotFoundError, ValidationAppError
from app.models.enums import ExportStatus, JobStatus, TagStatus
from app.models.export import ExportHistory
from app.models.tag_row import ImportedTagRow
from app.models.workbook import WorkbookJob
from app.services.workbook_service import get_job

# In-process guard against concurrent exports of the same job.
_export_locks: dict[str, threading.Lock] = {}
_locks_guard = threading.Lock()


def _job_lock(job_id: uuid.UUID) -> threading.Lock:
    with _locks_guard:
        return _export_locks.setdefault(str(job_id), threading.Lock())


def _output_filename(job: WorkbookJob) -> str:
    project = (job.project_code or "JOB").replace(" ", "")
    rev = (job.revision or "REV").replace(" ", "")
    stamp = datetime.now(UTC).strftime("%Y-%m-%d_%H%M%S")
    return f"{project}_{rev}_additional_info_{stamp}.xlsx"


def export_job(db: Session, job_id: uuid.UUID, user_id: uuid.UUID | None) -> ExportHistory:
    job = get_job(db, job_id)
    lock = _job_lock(job_id)
    if not lock.acquire(blocking=False):
        raise ConflictError("An export is already running for this job", code="EXPORT_IN_PROGRESS")
    try:
        return _do_export(db, job, user_id)
    finally:
        lock.release()


def _do_export(db: Session, job: WorkbookJob, user_id: uuid.UUID | None) -> ExportHistory:
    if not job.sheet_name or not job.additional_info_column_number:
        raise ValidationAppError("Job is not validated for export", code="JOB_NOT_VALIDATED")

    original = Path(job.original_file_path)
    if not original.exists():
        raise NotFoundError("Original workbook file is missing")

    rows = db.scalars(
        select(ImportedTagRow).where(
            ImportedTagRow.workbook_job_id == job.id,
            ImportedTagRow.status.in_(
                [TagStatus.COMPLETED.value, TagStatus.REVIEWED.value]
            ),
            ImportedTagRow.generated_additional_information.isnot(None),
        )
    ).all()

    export_dir = Path(settings.export_dir)
    export_dir.mkdir(parents=True, exist_ok=True)
    output_name = _output_filename(job)
    output_path = export_dir / output_name

    record = ExportHistory(
        workbook_job_id=job.id,
        output_filename=output_name,
        output_file_path=str(output_path),
        row_count_written=0,
        generated_by=user_id,
        status=ExportStatus.PENDING.value,
    )
    db.add(record)
    db.commit()
    db.refresh(record)

    try:
        # Load the ORIGINAL (never modified) and write a fresh copy.
        wb = load_workbook(filename=str(original))
        ws = wb[job.sheet_name]
        col = job.additional_info_column_number
        written = 0
        for r in rows:
            if r.sheet_name != job.sheet_name:
                # Only the detected sheet is exported in this version.
                continue
            ws.cell(row=r.excel_row_number, column=col).value = (
                r.generated_additional_information
            )
            written += 1

        # Atomic write: temp file in the same dir, then os.replace.
        fd, tmp_name = tempfile.mkstemp(suffix=".xlsx", dir=str(export_dir))
        os.close(fd)
        wb.save(tmp_name)
        wb.close()

        # Verify the workbook reopens before publishing it.
        verify = load_workbook(filename=tmp_name, read_only=True)
        verify.close()

        os.replace(tmp_name, output_path)
        digest = hashlib.sha256(output_path.read_bytes()).hexdigest()

        # Mark exported rows.
        for r in rows:
            if r.sheet_name == job.sheet_name:
                r.status = TagStatus.EXPORTED.value

        record.row_count_written = written
        record.file_hash = digest
        record.status = ExportStatus.SUCCESS.value
        job.status = JobStatus.EXPORTED.value
        db.commit()
        db.refresh(record)
        return record
    except Exception as exc:
        db.rollback()
        record = db.get(ExportHistory, record.id)
        if record is not None:
            record.status = ExportStatus.FAILED.value
            record.error_message = str(exc)[:2000]
            db.commit()
        raise


def list_exports(db: Session, job_id: uuid.UUID) -> list[ExportHistory]:
    return list(
        db.scalars(
            select(ExportHistory)
            .where(ExportHistory.workbook_job_id == job_id)
            .order_by(ExportHistory.generated_at.desc())
        ).all()
    )


def get_export(db: Session, export_id: uuid.UUID) -> ExportHistory:
    record = db.get(ExportHistory, export_id)
    if record is None:
        raise NotFoundError("Export not found")
    return record
